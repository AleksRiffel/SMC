from docxtpl import DocxTemplate
import openpyxl
from ..core.GeneralDict import *
from openpyxl import Workbook

def load_excel(excel_file_path: str) -> Workbook:
    """
    Загружает Excel-файл с данными.
    
    Args:
        excel_file_path (str): Путь к Excel-файлу.
    
    Returns:
        Workbook: Объект рабочей книги Excel.
    """
    wb = openpyxl.load_workbook(filename=excel_file_path)
    return wb

def extract_basic_info(wb: Workbook) -> dict:
    """
    Извлекает общую информацию о направлении, профиле и списке дисциплин.
    
    Args:
        wb (Workbook): Объект рабочей книги Excel.
    
    Returns:
        dict: Словарь с ключами:
            - "направление": Название направления.
            - "профиль": Название профиля.
            - "список дисциплин": Список дисциплин.
            - "кафедры": Список кафедр.
    """
    names = []
    names.append(wb['ПланСвод']["C6"].value)
    for i in range(7, 10000):
        if wb['ПланСвод']["C" + str(i)].value is None:
            if wb['ПланСвод']["C" + str(i + 1)].value is None:
                break
            continue
        if wb['ПланСвод']["C" + str(i - 1)].value is None or wb['ПланСвод']["C" + str(i)].font.bold:
            continue
        names.append(wb['ПланСвод']["C" + str(i)].value)
    
    namesKafedras = []
    for i in range(2, 10000):
        if wb['Кафедры']["C" + str(i)].value is None:
            break
        namesKafedras.append(wb['Кафедры']["C" + str(i)].value)
    
    return {
        "направление": str(wb['Титул']["B18"].value.split("Направление: ")[1]),
        "профиль": wb['Титул']["B19"].value,
        "список дисциплин": names,
        "кафедры": namesKafedras
    }

def extract_specific_info(wb: Workbook) -> dict:
    """
    Извлекает специфичную информацию из файла (например, форму обучения).
    
    Args:
        wb (Workbook): Объект рабочей книги Excel.
    
    Returns:
        dict: Словарь с ключами:
            - "форма обучения": Форма обучения.
    """
    return {
        "форма обучения": wb['Титул']["A31"].value.split("Форма обучения: ")[1].split(" ")[0].lower()
    }

def extract_competences(wb: Workbook, pickedname: str) -> list:
    """
    Извлекает коды компетенций для указанной дисциплины.
    
    Args:
        wb (Workbook): Объект рабочей книги Excel.
        pickedname (str): Название дисциплины.
    
    Returns:
        list: Список кодов компетенций.
    """
    codesCompitence = []
    for i in range(4, 10000):
        if wb['Компетенции(2)']["F" + str(i)].value is None:
            break
        if wb['Компетенции(2)']["F" + str(i)].value == pickedname:
            codesCompitence = wb['Компетенции(2)']["G" + str(i)].value.split("; ")
            break
    return codesCompitence

def fill_competences(wb: Workbook, codesCompitence: list) -> list:
    """
    Заполняет компетенции в GeneralDict на основе предоставленных кодов.
    
    Args:
        wb (Workbook): Объект рабочей книги Excel.
        codesCompitence (list): Список кодов компетенций.
    
    Returns:
        list: Список объектов Компетенция.
    """
    competences = []
    for i in range(3, 10000):
        if wb['Компетенции']["D" + str(i)].value is None:
            break
        if wb['Компетенции']["B" + str(i)].value is None:
            continue
        for j in codesCompitence:
            if wb['Компетенции']["B" + str(i)].value == j:
                competences.append(Компетенция(
                    код=j,
                    описание=wb['Компетенции']["D" + str(i)].value
                ))
                break
        if len(competences) == len(codesCompitence):
            break
    return competences

def fill_discipline_volume(wb: Workbook, pickedname: str) -> dict:
    """
    Заполняет объем дисциплины: часы, зачетные единицы, виды занятий.
    
    Args:
        wb (Workbook): Объект рабочей книги Excel.
        pickedname (str): Название дисциплины.
    
    Returns:
        dict: Словарь с ключами:
            - "часы": Общее количество часов.
            - "зачетные единицы": Количество зачетных единиц.
            - "виды занятий": Список объектов ХарактеристикиОбъемаДисциплины.
    """
    volume = []
    часы = ""
    зачетные_единицы = ""
    
    for i in range(7, 10000):
        if wb['План']["C" + str(i)].value == pickedname:
            часы = wb['План']["L" + str(i)].value
            зачетные_единицы = wb['План']["I" + str(i)].value
            
            # Определение семестров в предмете
            listsemestrs = []
            if wb['План']["D" + str(i)].value is not None:
                for number in list(str(wb['План']["D" + str(i)].value)):
                    listsemestrs.append(["экзамен", int(number)])
            if wb['План']["E" + str(i)].value is not None:
                for number in list(str(wb['План']["E" + str(i)].value)):
                    listsemestrs.append(["экзамен", int(number)])
            
            listsemestrs.sort(key=lambda x: x[1])
            
            for infosemestr in listsemestrs:
                infosemestr[1] = infosemestr[1]
                hoursdict = ХарактеристикиОбъемаДисциплины(
                    лекции=str(wb['План'].cell(row=i, column=17+(infosemestr[1]-1)*7+1).value) if wb['План'].cell(row=i, column=17+(infosemestr[1]-1)*7+1).value is not None else "-",
                    практика=str(wb['План'].cell(row=i, column=17+(infosemestr[1]-1)*7+3).value) if wb['План'].cell(row=i, column=17+(infosemestr[1]-1)*7+3).value is not None else "-",
                    лабраб=str(wb['План'].cell(row=i, column=17+(infosemestr[1]-1)*7+2).value) if wb['План'].cell(row=i, column=17+(infosemestr[1]-1)*7+2).value is not None else "-", 
                    самраб=str(wb['План'].cell(row=i, column=17+(infosemestr[1]-1)*7+5).value) if wb['План'].cell(row=i, column=17+(infosemestr[1]-1)*7+5).value is not None else "-",
                    контроль=str(wb['План'].cell(row=i, column=17+(infosemestr[1]-1)*7+6).value) if wb['План'].cell(row=i, column=17+(infosemestr[1]-1)*7+6).value is not None else "-",
                    аттестация=infosemestr[0],
                    семестр=str(infosemestr[1]),
                    курс=str(infosemestr[1]//2 + infosemestr[1]%2)
                )
                volume.append(hoursdict)
            break
    
    return {
        "часы": часы,
        "зачетные единицы": зачетные_единицы,
        "виды занятий": volume
    }

def check_coursework(wb: Workbook, pickedname: str) -> bool:
    """
    Проверяет наличие курсовой работы для указанной дисциплины.
    
    Args:
        wb (Workbook): Объект рабочей книги Excel.
        pickedname (str): Название дисциплины.
    
    Returns:
        bool: True, если курсовая работа есть, иначе False.
    """
    isworks = False
    for i in range(2, 1000):
        if wb['Курсовые']["A" + str(i)].value == pickedname + " ":
            isworks = True
            break
    return isworks

def generate_documents(
        content: GeneralDictModel,
        template_rp: str = "шаблонРП.docx",
        template_a: str = "шаблонА.docx",
        template_fos: str = "шаблонФОС.docx",
        output_rp: str = "generated_RP.docx", 
        output_a: str = "generated_A.docx",
        output_fos: str = "generated_fos.docx") -> None:
    """
    Генерирует документы по шаблонам на основе предоставленного контента.
    
    Args:
        content (GeneralDictModel): Данные для заполнения шаблонов.
        template_rp (str): Путь к шаблону РП.
        template_a (str): Путь к шаблону А.
        template_fos (str): Путь к шаблону ФОС.
        output_rp (str): Путь для сохранения РП.
        output_a (str): Путь для сохранения А.
        output_fos (str): Путь для сохранения ФОС.
    """
    doc = DocxTemplate(template_rp)
    doc.render(content.model_dump())
    doc.save(output_rp)
    
    doc = DocxTemplate(template_a)
    doc.render(content.model_dump())
    doc.save(output_a)
    
    doc = DocxTemplate(template_fos)
    doc.render(content.model_dump())
    doc.save(output_fos)